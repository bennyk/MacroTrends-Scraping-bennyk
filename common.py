from openpyxl import Workbook, worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import re
import pandas as pd
from openpyxl import Workbook, worksheet
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment

from itertools import tee
from typing import List, Dict


def pairwise(iterable):
    "s -> (s0,s1), (s1,s2), (s2, s3), ..."
    a, b = tee(iterable)
    next(b, None)
    return zip(a, b)


def isnumeric(a: str):
    return re.match(r'[.0-9]+$', a)


def get_driver(options):
    driver = webdriver.Firefox(executable_path=r'C:\Users\benny\PyCharmProjects\MacroTrends-Scraping\geckodriver.exe',
                               options=options)
    return driver


def get_options():
    options = Options()
    options.binary_location = r'C:\Users\benny\AppData\Local\Mozilla Firefox\firefox.exe'
    return options


def current_URL(url_path):
    options = get_options()
    driver = get_driver(options)
    driver.get(url_path)
    box = driver.find_element(By.CSS_SELECTOR, ".js-typeahead")
    ticker = 'intc'
    box.send_keys(ticker)
    time.sleep(1)
    box.send_keys(Keys.DOWN, Keys.RETURN)
    time.sleep(1)
    current_url = driver.current_url
    time.sleep(4)
    driver.quit()
    return current_url


def parse_grid(driver, fin_url):
    driver.get(fin_url)
    driver.set_window_size(2000, 2000)
    time.sleep(4)
    arrow = driver.find_element(By.CSS_SELECTOR, ".jqx-icon-arrow-right")
    webdriver.ActionChains(driver).click_and_hold(arrow).perform()
    # webdriver.ActionChains(driver).click_and_hold(arrow).move_by_offset(-1500, 0).release().perform()
    time.sleep(4)
    column_grid = driver.find_element(By.CSS_SELECTOR, "#columntablejqxgrid").text

    return parse_content(driver, column_grid)


def parse_content(driver, col_grid):
    # prune it for $ and comma
    content_grid = driver.find_element(By.CSS_SELECTOR, "#contenttablejqxgrid").text
    content_grid = content_grid.replace("$", "").replace(",", "")
    content_grid = content_grid.splitlines()
    data_dict = {'Years': col_grid.split('\n')[1:]}
    last_key = None
    for s in content_grid:
        if not (isnumeric(s) or s == '-' or s.startswith('-')):
            last_key = s
        else:
            if last_key is None:
                continue  # Skip data if no key is available

            try:
                ticker = float(s)
                data_dict.setdefault(last_key, []).append(ticker)
            except ValueError:
                data_dict.setdefault(last_key, []).append(s)
    return data_dict


def curly_brace(my_str, open_list=None, close_list=None):
    # Function to check parentheses

    # TODO Could not solve
    # for a in re.finditer(r'\[([^]]+)\]\[([^]]*)\]', g):

    if close_list is None:
        close_list = ["}"]
    if open_list is None:
        open_list = ["{"]

    def _check(str):
        # https://www.geeksforgeeks.org/check-for-balanced-parentheses-in-python/
        stack = []
        string = ''
        level = 0
        for c in str:
            if c in open_list:
                stack.append(c)
                level += 1
                if level == 1:
                    string = ''
            elif c in close_list:
                pos = close_list.index(c)
                if ((len(stack) > 0) and
                        (open_list[pos] == stack[len(stack) - 1])):
                    stack.pop()
                    level -= 1
                    if level == 0:
                        # print(string)
                        yield string
                else:
                    return "Unbalanced"
            else:
                string += c
        if len(stack) == 0:
            return "Balanced"
        else:
            return "Unbalanced"

    return _check(my_str)


class MyVar:
    def __init__(self):
        self.var = 1

    def gen_var(self):
        self.var += 1
        return self.var


class Clipboard:
    def __init__(self):
        self.wb = Workbook()
        self.excel_file = 'revenue_and_profit.xlsx'

        # removing initial sheet
        ws = self.wb.active
        self.wb.remove(ws)

    def write_excel(self, title, df):
        ws = self.wb.create_sheet(title)

        # Add the column headers
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    def write_dict_to_excel(self, title, data_dict):
        df = pd.DataFrame(data_dict)
        df = df.replace('-', 0)
        df = df.sort_values(by='Years', ascending=True)
        ws = self.wb.create_sheet(title)

        # Add the column headers
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

    def save(self):
        # Save the Excel file
        self.wb.save(self.excel_file)

    def insert_newsheet(self, new_sheetname: str, col_tuples: List[Dict]):
        ws = self.wb.create_sheet(new_sheetname)
        new_sheet_index = len(self.wb.sheetnames) - 1
        self.wb.active = new_sheet_index

        # Iterate to ~60 rows
        # len 58
        for i in range(1, 60):
            ws.cell(column=1, row=i, value='=Income!{letter}{index}'.format(index=i, letter=column_to_letter(1)))

        for t in col_tuples:
            # if t['text'] != '':
            #     if 'ignore' in t:
            #         continue
            for i in range(1, 60):
                ws.cell(column=t['idx'], row=i, value='={table}!{letter}{index}'
                        .format(index=i, letter=column_to_letter(t['ref']), table=t['table']))
        return ws

    def insert_chart(self, ws, start, end, new_sheetname):

        # TODO insert start result col and end result col
        for j in range(1, end):
            ws.column_dimensions[column_to_letter(j)].width = 10
            ws.cell(column=j, row=1).alignment = Alignment(wrapText=True)

        # Graph data
        chart = LineChart()
        for i in range(start, end):
            letter = column_to_letter(i)
            data = Reference(ws, range_string=f'{new_sheetname}!{letter}1:{letter}60')
            chart.add_data(data, titles_from_data=True)

        category = Reference(ws, range_string=f'{new_sheetname}!A2:A60')
        chart.set_categories(category)
        ws.add_chart(chart, 'D3')

    def insert_income(self):
        var = MyVar()
        rev_col = var.gen_var()
        gross_col = var.gen_var()
        op_col = var.gen_var()
        net_col = var.gen_var()
        ocf_col = var.gen_var()
        capex_col = var.gen_var()
        fcf_col = var.gen_var()
        rev_growth_col = var.gen_var()
        gross_margin_col = var.gen_var()
        op_margin_col = var.gen_var()
        net_margin_col = var.gen_var()
        ocf_margin_col = var.gen_var()
        fcf_margin_col = var.gen_var()

        # Column tuple for reference, new index, and text ref in Excel.
        # TODO Industry specific ration CFO / Capex
        # https://www.investopedia.com/terms/c/capitalexpenditure.asp
        col_tuples = [{'ref': 2, 'idx': rev_col, 'text': 'Revenue', 'table': 'Income'},
                      {'ref': 4, 'idx': gross_col, 'text': 'Gross', 'table': 'Income'},
                      {'ref': 9, 'idx': op_col, 'text': 'Operating', 'table': 'Income'},
                      {'ref': 17, 'idx': net_col, 'text': 'Net', 'table': 'Income'},
                      {'ref': 11, 'idx': ocf_col, 'text': 'OCF', 'table': 'Cash'},
                      # Ignoring flag in Net change in PPE
                      {'ref': 12, 'idx': capex_col, 'text': 'Capex margin', 'table': 'Cash', 'ignore': True},
                      ]
        new_sheetname = 'Margins'
        ws = self.insert_newsheet(new_sheetname, col_tuples)

        ws.cell(column=rev_growth_col, row=1, value='Revenue growth %')
        for i in range(1, 60):
            if (i-4) > 1:
                ws.cell(column=rev_growth_col, row=i, value='=({letter}{index}-{letter}{index2})/{letter}{index2}'
                        .format(index=i, index2=i-4, letter=column_to_letter(rev_col)))
                ws['{letter}{index}'.format(
                    index=i, letter=column_to_letter(rev_growth_col))].number_format = '0%'

        for col, margin_col, text in [
            (gross_col, gross_margin_col, 'Gross margin'),
            (op_col, op_margin_col, 'Op margin'),
            (net_col, net_margin_col, 'Net margin'),
            (ocf_col, ocf_margin_col, 'OCF margin'),
        ]:
            ws.cell(column=margin_col, row=1, value=text)
            for i in range(2, 60):
                ws.cell(column=margin_col, row=i,
                        value='=${col}{index}/{rev}{index}'
                        .format(index=i,
                                rev=column_to_letter(rev_col),
                                col=column_to_letter(col), ))
                ws['{}{}'.format(column_to_letter(margin_col), i)].number_format = '0%'

        ws.cell(column=fcf_col, row=1, value='FCF')
        ws.cell(column=fcf_margin_col, row=1, value='FCF margin')
        for i in range(2, 60):
            ws.cell(column=fcf_col, row=i, value='=({ocf}{index}+{capex}{index})'
                    .format(index=i,
                            ocf=column_to_letter(ocf_col),
                            capex=column_to_letter(capex_col), ))
            ws['{}{}'.format(column_to_letter(fcf_col), i)].number_format = '0.0'

            ws.cell(column=fcf_margin_col, row=i, value='={fcf}{index}/{rev}{index}'
                    .format(index=i,
                            fcf=column_to_letter(fcf_col),
                            rev=column_to_letter(rev_col), ))
            ws['{}{}'.format(column_to_letter(fcf_margin_col), i)].number_format = '0%'

        self.insert_chart(ws, rev_growth_col, fcf_margin_col+1, new_sheetname)

    def insert_debt(self):
        var = MyVar()
        cash_col = var.gen_var()
        lt_assets_col = var.gen_var()
        total_current_liab_col = var.gen_var()
        lt_liabilities_col = var.gen_var()
        equity_col = var.gen_var()
        lt_liabilities_to_assets_col = var.gen_var()
        net_debt_to_equity_col = var.gen_var()
        cash_to_short_term_col = var.gen_var()

        col_tuples = [
            # Cash
            {'ref': 2, 'idx': cash_col, 'text': 'LT asset', 'table': 'Balance'},
            # LT assets
            {'ref': 12, 'idx': lt_assets_col, 'text': 'LT asset', 'table': 'Balance'},
            # ST debt
            {'ref': 14, 'idx': total_current_liab_col, 'text': 'LT debt', 'table': 'Balance'},
            # LT liabilities
            {'ref': 17, 'idx': lt_liabilities_col, 'text': 'LT liabilities', 'table': 'Balance'},
            # Shareholder equity
            {'ref': 23, 'idx': equity_col, 'text': 'Equity', 'table': 'Balance'},
        ]
        new_sheetname = 'Debt'
        ws = self.insert_newsheet(new_sheetname, col_tuples)

        ws.cell(column=lt_liabilities_to_assets_col, row=1, value='Long term liabilities to assets')
        ws.cell(column=net_debt_to_equity_col, row=1, value='Net debt to equity')
        ws.cell(column=cash_to_short_term_col, row=1, value='Cash to short-term borrowing')
        for i in range(2, 60):
            ws.cell(column=lt_liabilities_to_assets_col, row=i, value='=${lt_liabilities}{index}/${lt_assets}{index}'
                    .format(index=i,
                            lt_liabilities=column_to_letter(lt_liabilities_col),
                            lt_assets=column_to_letter(lt_assets_col)))
            ws['{}{}'.format(column_to_letter(lt_liabilities_to_assets_col), i)].number_format = '0%'

            ws.cell(column=net_debt_to_equity_col, row=i,
                    value='=(${total_lt_liab}{index}+${total_current_liab}{index}-${cash}{index})/${equity}{index}'
                    .format(index=i,
                            cash=column_to_letter(cash_col),
                            total_current_liab=column_to_letter(total_current_liab_col),
                            total_lt_liab=column_to_letter(lt_liabilities_col),
                            equity=column_to_letter(equity_col), ))
            ws['{}{}'.format(column_to_letter(net_debt_to_equity_col), i)].number_format = '0%'

            ws.cell(column=cash_to_short_term_col, row=i,
                    value='=${cash}{index}/${total_current_liab}{index}'
                    .format(index=i,
                            cash=column_to_letter(cash_col),
                            total_current_liab=column_to_letter(total_current_liab_col), ))
            ws['{}{}'.format(column_to_letter(cash_to_short_term_col), i)].number_format = '0.00'

        self.insert_chart(ws, lt_liabilities_to_assets_col, cash_to_short_term_col + 1,
                          new_sheetname)

    def insert_returns(self):
        var = MyVar()
        op_income_col = var.gen_var()
        tax_col = var.gen_var()
        net_income_col = var.gen_var()
        asset_col = var.gen_var()
        debt_col = var.gen_var()
        equity_col = var.gen_var()
        cash_from_investing_col = var.gen_var()
        cash_from_financing_col = var.gen_var()
        roe_col = var.gen_var()
        roa_col = var.gen_var()
        roic_col = var.gen_var()

        col_tuples = [
            # Op income
            {'ref': 9, 'idx': op_income_col, 'text': 'Op income', 'table': 'Income'},
            # Tax
            {'ref': 12, 'idx': tax_col, 'text': 'Tax expense', 'table': 'Income'},
            # Net income
            {'ref': 17, 'idx': net_income_col, 'text': 'Net income', 'table': 'Income'},
            # Assets
            {'ref': 13, 'idx': asset_col, 'text': "Shareholders' equity", 'table': 'Balance'},
            # Debt
            {'ref': 15, 'idx': debt_col, 'text': "Debt", 'table': 'Balance'},
            # Equity
            {'ref': 23, 'idx': equity_col, 'text': "Equity", 'table': 'Balance'},
            # Cash from investing
            {'ref': 19, 'idx': cash_from_investing_col, 'text': "Cash from investing", 'table': 'Cash'},
            # Cash from financing
            {'ref': 27, 'idx': cash_from_financing_col, 'text': "Cash from financing", 'table': 'Cash'},
        ]

        new_sheetname = 'Returns'
        ws = self.insert_newsheet(new_sheetname, col_tuples)

        ws.cell(column=roe_col, row=1, value='Return on Equity')
        ws.cell(column=roa_col, row=1, value='Return on Asset')
        ws.cell(column=roic_col, row=1, value='Return on Invested Capital')
        for i in range(5, 60):
            ws.cell(
                column=roe_col, row=i,
                value='=SUM(${net_income}{index1}:{net_income}{index2})/AVERAGE(${equity}{index1}:${equity}{index2})'
                .format(index1=i - 3,
                        index2=i,
                        net_income=column_to_letter(net_income_col),
                        equity=column_to_letter(equity_col)))
            ws['{}{}'.format(column_to_letter(roe_col), i)].number_format = '0%'

            ws.cell(
                column=roa_col, row=i,
                value='=SUM(${net_income}{index1}:{net_income}{index2})/AVERAGE(${asset}{index1}:${asset}{index2})'
                .format(index1=i - 3,
                        index2=i,
                        net_income=column_to_letter(net_income_col),
                        asset=column_to_letter(asset_col)))
            ws['{}{}'.format(column_to_letter(roa_col), i)].number_format = '0%'

            ws.cell(
                column=roic_col, row=i,
                value='=(SUM(${op_income}{index1}:{op_income}{index2})'
                      '- SUM(${tax}{index1}:{tax}{index2}))'
                      '/ (AVERAGE(${debt}{index1}:${debt}{index2})'
                      '+ AVERAGE(${equity}{index1}:${equity}{index2})'
                      '+ AVERAGE(${cash_from_investing}{index1}:${cash_from_investing}{index2})'
                      '+ AVERAGE(${cash_from_financing}{index1}:${cash_from_financing}{index2}))'
                .format(index1=i - 3,
                        index2=i,
                        op_income=column_to_letter(op_income_col),
                        tax=column_to_letter(tax_col),
                        debt=column_to_letter(debt_col),
                        equity=column_to_letter(equity_col),
                        cash_from_investing=column_to_letter(cash_from_investing_col),
                        cash_from_financing=column_to_letter(cash_from_financing_col), ))
            ws['{}{}'.format(column_to_letter(roic_col), i)].number_format = '0%'

        self.insert_chart(ws, roe_col, roic_col + 1, new_sheetname)

    def insert_eps(self):
        var = MyVar()
        eps_col = var.gen_var()
        annual_eps_col = var.gen_var()
        col_tuples = [
            # EPS
            {'ref': 23, 'idx': eps_col, 'text': 'EPS', 'table': 'Income'},
        ]
        new_sheetname = 'EPS'
        ws = self.insert_newsheet(new_sheetname, col_tuples)

        ws.cell(column=annual_eps_col, row=1, value='Annual EPS')
        for i in range(5, 60):
            ws.cell(column=annual_eps_col, row=i,
                    value='=SUM(${eps}{index1}:{eps}{index2})'
                    .format(index1=i - 3,
                            index2=i,
                            eps=column_to_letter(eps_col), ))
            ws['{}{}'.format(column_to_letter(annual_eps_col), i)].number_format = '0.00'

        self.insert_chart(ws, annual_eps_col, annual_eps_col + 1, new_sheetname)

    def insert_shares_out(self):
        var = MyVar()
        shares_out_col = var.gen_var()
        col_tuples = [
            # Shares outstanding
            {'ref': 21, 'idx': shares_out_col, 'text': 'Shares outstanding', 'table': 'Income'},
        ]
        new_sheetname = 'Shares'
        ws = self.insert_newsheet(new_sheetname, col_tuples)
        self.insert_chart(ws, shares_out_col, shares_out_col + 1, new_sheetname)


def old_run_main():
    main_url_path = 'https://www.macrotrends.net/'
    current_url = current_URL(main_url_path)

    if "stocks" in current_url:
        # Check if the data in the ticker is available
        url_parts = current_url.split("/", 10)
        url_path = main_url_path + "stocks/charts/" + url_parts[5] + "/" + url_parts[6] + "/"
        driver = get_driver(get_options())
        # financial-statements
        fin_url_path = url_path + "financial-statements"
        driver.get(fin_url_path)
        if driver.find_elements(
                By.CSS_SELECTOR,
                "div.jqx-grid-column-header:nth-child(1) > div:nth-child(1) > div:nth-child(1) > span:nth-child(1)"):
            clip = Clipboard()

            income_url = url_path + "income-statement"
            data_dict = parse_grid(driver, income_url)
            clip.write_excel('Income', data_dict)

            balance_url = url_path + "balance-sheet"
            data_dict = parse_grid(driver, balance_url)
            clip.write_excel('Balance', data_dict)

            cash_url = url_path + "cash-flow-statement"
            data_dict = parse_grid(driver, cash_url)
            clip.write_excel('Cash', data_dict)

            clip.save()
            driver.quit()


def column_to_letter(column_int):
    # https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
    start_index = 1  # It can start either at 0 or at 1
    letter = ''
    while column_int > 25 + start_index:
        letter += chr(65 + int((column_int - start_index) / 26) - 1)
        column_int = column_int - (int((column_int - start_index) / 26)) * 26
    letter += chr(65 - start_index + (int(column_int)))
    return letter


def letter_to_column(name):
    # https://cwestblog.com/2013/09/13/python-convert-excel-column-name-to-number/
    pow = 1
    column_int = 0
    for letter in name[::-1]:
        column_int += (int(letter, 36) - 9) * pow
        pow *= 26
    return column_int
